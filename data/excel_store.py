from datetime import date
from typing import Any, Dict, List, Optional, Tuple
from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from config.settings import EXCEL_PATH

# Load workbook and sheets
try:
    wb: Workbook = load_workbook(EXCEL_PATH)
except Exception as e:
    raise Exception(f"ERROR: Could not load {EXCEL_PATH}: {e}")

SHEET_DB: Worksheet = wb["dB"]
SHEET_PM: Worksheet = wb["Product Master"]
SHEET_INV: Worksheet = wb["Inventory"]
SHEET_OBS: Worksheet = wb["Non Stock Items"]

if "Transfer Log" not in wb.sheetnames:
    wb.create_sheet("Transfer Log")
SHEET_TL: Worksheet = wb["Transfer Log"]


def get_all_products() -> List[str]:
    products: List[str] = []
    for row in SHEET_PM.iter_rows(min_row=2, values_only=True):
        name = row[1]
        if name:
            products.append(str(name))
    return sorted(products)


def get_all_product_groups() -> List[str]:
    groups = set()
    for row in SHEET_PM.iter_rows(min_row=2, values_only=True):
        group = row[2]
        if group:
            groups.add(str(group))
    return sorted(groups)


def get_product_group_map() -> Dict[str, Optional[str]]:
    """Return a mapping of product name -> product group from the Product Master sheet."""
    mapping: Dict[str, Optional[str]] = {}
    for row in SHEET_PM.iter_rows(min_row=2, values_only=True):
        name = row[1]
        group = row[2] if len(row) > 2 else None
        if name:
            mapping[str(name)] = str(group) if group is not None else None
    return mapping


def get_product_group(product_name: str) -> Optional[str]:
    """Return the product group for a single product, or None if unknown."""
    for row in SHEET_PM.iter_rows(min_row=2, values_only=True):
        name = row[1]
        group = row[2] if len(row) > 2 else None
        if name == product_name:
            return str(group) if group is not None else None
    return None


def get_min_critical(product_name: str) -> Tuple[Optional[float], Optional[float]]:
    for row in SHEET_PM.iter_rows(min_row=2, values_only=True):
        if row[1] == product_name:
            return row[5], row[6]
    return (None, None)


def calculate_balance(product_name: str) -> float:
    balance: float = 0.0
    for row in SHEET_DB.iter_rows(min_row=2, values_only=True):
        prod = row[1]
        qty = row[4]
        stock_type = row[10]

        if prod == product_name:
            try:
                qty = int(float(qty))
            except Exception:
                qty = 0

            if stock_type == "Stock In":
                balance += qty
            elif stock_type == "Stock Out":
                balance -= qty

    return balance


def calculate_all_balances() -> Dict[str, Dict[str, float]]:
    balances: Dict[str, Dict[str, float]] = {}
    for row in SHEET_DB.iter_rows(min_row=2, values_only=True):
        product = row[1]
        qty = row[4]
        stock_type = row[10]

        if product is None:
            continue

        if product not in balances:
            balances[product] = {"in": 0.0, "out": 0.0, "balance": 0.0}

        try:
            qty = int(float(qty))
        except Exception:
            qty = 0

        if stock_type == "Stock In":
            balances[product]["in"] += qty
        elif stock_type == "Stock Out":
            balances[product]["out"] += qty

    for product in balances:
        balances[product]["balance"] = balances[product]["in"] - balances[product]["out"]

    return balances


def get_total_stock_in_out() -> Tuple[float, float]:
    total_in: float = 0.0
    total_out: float = 0.0
    for row in SHEET_DB.iter_rows(min_row=2, values_only=True):
        qty = row[4]
        stock_type = row[10]
        try:
            qty = int(float(qty))
        except Exception:
            qty = 0
        if stock_type == "Stock In":
            total_in += qty
        elif stock_type == "Stock Out":
            total_out += qty
    return total_in, total_out


def log_movement(stock_type: str, product: str, serial: Optional[str], model: Optional[str], qty: float, dept: Optional[str], pcname: Optional[str]) -> None:
    next_row = SHEET_DB.max_row + 1
    SHEET_DB.cell(next_row, 1).value = str(date.today())
    SHEET_DB.cell(next_row, 2).value = product
    SHEET_DB.cell(next_row, 3).value = serial
    SHEET_DB.cell(next_row, 4).value = model
    SHEET_DB.cell(next_row, 5).value = qty
    SHEET_DB.cell(next_row, 9).value = dept
    SHEET_DB.cell(next_row, 10).value = pcname
    SHEET_DB.cell(next_row, 11).value = stock_type
    wb.save(EXCEL_PATH)


def get_inventory_summary_rows() -> List[Any]:
    summary_rows: List[Any] = []
    for row in SHEET_INV.iter_rows(min_row=1, values_only=False):
        values = [cell.value for cell in row]
        if str(values[0]).strip() == "Product":
            continue
        product = values[0]
        qty_in = values[1] if len(values) > 1 else None
        qty_out = values[2] if len(values) > 2 else None
        if product and isinstance(qty_in, (int, float)) and isinstance(qty_out, (int, float)):
            summary_rows.append(row)
    return summary_rows


def count_obsolete() -> int:
    count: int = 0
    for row in SHEET_OBS.iter_rows(min_row=2, values_only=True):
        if row[0]:
            count += 1
    return count


def update_inventory_summary() -> None:
    summary_rows = get_inventory_summary_rows()
    product_totals: Dict[str, Dict[str, float]] = {}
    for row in SHEET_DB.iter_rows(min_row=2, values_only=True):
        prod = row[1]
        qty = row[4]
        stock_type = row[10]
        if prod:
            if prod not in product_totals:
                product_totals[prod] = {"in": 0.0, "out": 0.0}
            try:
                qty = int(float(qty))
            except Exception:
                qty = 0
            if stock_type == "Stock In":
                product_totals[prod]["in"] += qty
            elif stock_type == "Stock Out":
                product_totals[prod]["out"] += qty

    for row_cells in summary_rows:
        product_cell = row_cells[0]
        product_name = product_cell.value
        if not product_name:
            continue
        qty_in_cell = row_cells[1]
        qty_out_cell = row_cells[2]
        bal_cell = row_cells[3]
        status_cell = row_cells[6]
        totals = product_totals.get(product_name, {"in": 0.0, "out": 0.0})
        qty_in = totals["in"]
        qty_out = totals["out"]
        balance = qty_in - qty_out
        qty_in_cell.value = qty_in
        qty_out_cell.value = qty_out
        bal_cell.value = balance
        min_level, critical_level = get_min_critical(product_name)
        try:
            from services.ui_helpers import compute_status
            status_text = compute_status(balance, min_level, critical_level)
        except Exception:
            status_text = ""
        status_cell.value = status_text

    wb.save(EXCEL_PATH)


def get_stock_on_hand_table() -> List[Dict[str, Any]]:
    stock_list: List[Dict[str, Any]] = []
    for row in SHEET_INV.iter_rows(min_row=1, values_only=True):
        if row[0] and isinstance(row[1], (int, float)) and isinstance(row[2], (int, float)):
            product = row[0]
            qty_in = row[1]
            qty_out = row[2]
            balance = row[3]
            low = row[4] if len(row) > 4 else None
            critical = row[5] if len(row) > 5 else None
            status = row[6] if len(row) > 6 else None
            stock_list.append({
                "product": product,
                "qty_in": qty_in,
                "qty_out": qty_out,
                "balance": balance,
                "low": low,
                "critical": critical,
                "status": status
            })
    return stock_list
